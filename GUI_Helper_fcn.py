#!/usr/bin/env python
# coding: utf-8

# In[1]:


import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from scipy import ndimage
from scipy.ndimage.morphology import binary_fill_holes
from scipy.fft import fft,rfft,rfftfreq, irfft
import scipy
import seaborn as sns
import cv2
from sklearn.cluster import DBSCAN
from scipy.optimize import curve_fit
from skimage import measure
from skimage import morphology
from skimage import io
from skimage.segmentation import flood
from skimage.measure import regionprops
from skimage.morphology import binary_erosion, binary_dilation
import skimage
import pandas as pd
import requests
import queue 
import copy
import pickle
from mpl_interactions import hyperslicer
from openpyxl import load_workbook
from pylab import *

get_ipython().run_line_magic('matplotlib', 'ipympl')


# In[2]:


def setGlobal_GUI(fpsVal,scaleVal, name, aspectVal):
    #set global parameters
    #input:fpsVal: frame per second
    #      scaleVal: pixel distance scale
    #      name:sample name
    #      aspectVal:aspect of image height width
    global file_prefix
    file_prefix = name
    global fps
    fps = fpsVal
    global scale
    scale = scaleVal
    global sampleIndex
    sampleIndex = name
    global aspect
    aspect = aspectVal


# In[3]:


def videoIn(inputname):
    #import video and perform basic noise reduction
    #input:filename
    #output:image array im (true false 3D array)
    im = io.imread(inputname)  #all 0 1 entry
    im = im.astype(bool)  #convert to true false entry

    
    for i in range(im.shape[0]):   #for each image slice
        slice = ndimage.binary_closing(im[i,:,:], iterations = 1)  #binary closing of image to remove small holes
        slice = morphology.remove_small_objects(slice,
                                                    min_size= 100, connectivity=4)  #remove objects smaller than min size, connectivity
                                                                                    #defines the neighborhood of a pixel
        slice = ~slice  #invert true false entry
        slice = ndimage.binary_closing(slice, iterations = 1)
        im[i, :, :] = morphology.remove_small_objects(slice,
                                                    min_size= 100, connectivity=4)  #repeat same process and update image array
    return im


# In[4]:


def STmap(im, light_time = None,light_time_ob = None, light_distance = None,
          c_low = None, c_high = None,d_min = None, d_max = None,t_min = None, t_max = None, 
          z_score = False, normalization = False, saveFlag = False):
    #compute the spatial temporal map
    #input: image matrix im (3D:slice x row number x column number)
    #       the time for light stimulation: light_time
    #       the observation time period after stimulation: light_time_ob
    #       the distance of stimulation: light_distance
    #       colormap lower limit: c_low (default: None)
    #       colormap upper limit: c_high (default: None)
    #       heatmap visualization distance lower limit: d_min (default:None)
    #       heatmap visualization distance upper limit: d_max (default:None)
    #       heatmap visualization time lower limit: t_min (default:None)
    #       heatmap visualization time upper limit: t_max (default: None)
    #       whether to output zscore matrix and plot: z_score (default: None)
    #       whether to output normalized matrix with entry 0~1 and plot: normalization (default: None)
    #       whether to save the figure: saveFlag (default: False)
    #output: colormap of STmap, return STmap matrix
    im_s_r = np.sum(~im, axis = 1)[:,1:-1]/scale #sum up along each column in each image slice
                                 #turn row data in every column in each slice into one number: thickness/width of tissue
                                 #dimension:slice number x column number
    #filter out frames with low quality (noise frames)
    im_s = []
    for i in range(im_s_r.shape[0]):
        row_quality = 1
        for j in range(im_s_r.shape[1]):
            if im_s_r[i][j] < 0.1:   #filter out regions with noises
                row_quality = 0
        if row_quality == 1:
            im_s.append(im_s_r[i])
    im_s = np.array(im_s)
    
    if z_score:  #if compute z score map
        im_z = np.zeros(im_s.shape)  #initialize im z score matrix
        for i in range(im_s.shape[1]):   #range through each column
            im_z[:, i] = scipy.stats.zscore(im_s[:, i])
        figType = "Z_score_STmap"
        ST_map_visual(im_z,light_time,light_time_ob, light_distance, c_low, c_high,d_min, d_max,t_min, t_max,saveFlag, figType)
        return im_z
    elif normalization:  #if compute normalized map
        im_p = np.zeros(im_s.shape)
        for i in range(im_s.shape[1]):
            im_p[:, i] = im_s[:, i]/np.max(im_s[:,i])
        figType = "Normalized_STmap"
        ST_map_visual(im_p,light_time,light_time_ob, light_distance,c_low, c_high,d_min, d_max,t_min, t_max,saveFlag, figType)
        return im_p
    else:
        figType = "Raw_diameter_STmap"
        ST_map_visual(im_s,light_time,light_time_ob, light_distance,c_low, c_high,d_min, d_max,t_min, t_max,saveFlag, figType)
        return im_s


# In[5]:


def ST_map_visual(im,light_time,light_time_ob, light_distance,c_low, c_high,d_min, d_max,t_min, t_max,saveFlag, figType):
    #visualization for ST maps
    #input: image matrix im (3D:slice x row number x column number)
    #       the time for light stimulation: light_time
    #       the observation time period after stimulation: light_time_ob
    #       colormap lower limit: c_low (default: None)
    #       colormap upper limit: c_high (default: None)
    #       heatmap visualization distance lower limit: d_min (default:None)
    #       heatmap visualization distance upper limit: d_max (default:None)
    #       heatmap visualization time lower limit: t_min (default:None)
    #       heatmap visualization time upper limit: t_max (default: None)
    #       whether to save the figure: saveFlag
    #       figure type among three: figType
    #output: colormap of STmap
    #plt.close("all")
    plt.figure()
    plt.imshow(im, aspect = aspect, extent=[0,shape(im)[1]/scale,shape(im)[0]/fps,0])
    if c_low is None:  
        c_low = im.min()
    if c_high is None:
        c_high = im.max()
    if d_min is None:  
        d_min = 0
    if d_max is None:
        d_max = shape(im)[1]/scale
    if t_min is None:  
        t_min = 0
    if t_max is None:
        t_max = shape(im)[0]/fps              
    plt.set_cmap('jet')
    plt.colorbar()
    plt.clim([c_low, c_high])
    plt.xlim([d_min,d_max])
    plt.ylim([t_max,t_min])
    plt.xlabel("Distance(cm)")
    plt.ylabel("Time(s)")
    plt.show()
    if light_time != None:
        for i in light_time:
            plt.axhline(y = i, color='white', linestyle='-')
            plt.axhline(y = i + light_time_ob, color='white', linestyle='--')
    if light_time !=None:
        plt.axvline(x = light_distance/scale, color='white', linestyle='-')
    if saveFlag == True:
        plt.savefig("STmap/{}_{}.pdf".format(figType,sampleIndex))


# In[6]:


def Dmap3D(im):
    #compute 3D plot for diameter, x axis represent time, y axis represent distance
    #input: 2D image array: im
    fig = plt.figure()
    ax = fig.add_subplot(projection='3d')

    X = np.linspace(0, shape(im_s)[1]/scale, num=shape(im_s)[1]) #distance
    Y = np.linspace(0, shape(im_s)[0]/fps, num=shape(im_s)[0]) #time
    X, Y = np.meshgrid(X, Y)
    # Plot the surface.
    surf = ax.plot_surface(X, Y, im, cmap=cm.coolwarm,rstride=1, cstride=1,
                           linewidth=0, antialiased=True)
    fig.colorbar(surf, shrink=0.5, aspect=5)
    ax.set_xlabel("Distance(cm)")
    ax.set_ylabel("Time(s)")
    plt.show()


# In[7]:


def DmapFreqFFTVis(im_s, position):
    #compute 2D diameter plot at specific distance and perform fourier transform to extract frequency information
    #input: 2D image array: im_s
    #       the position index at distance map: position
    Tarray = np.linspace(0, shape(im_s)[0]/fps, num=shape(im_s)[0]) #time
    Darray = im_s[:,position] #get diameter at specific distance 
    Darray = ndimage.gaussian_filter1d(Darray, 50)  #smooth the curve
    Darrayfreqy = rfft(Darray)
    Darrayfreqx = rfftfreq(shape(im_s)[0],1/fps)  #perform fourier transform
    DarrayInv = irfft(Darrayfreqy, Darray.shape[0])   #perform inverse fourier transform to see if features are captured and preserved
    plt.figure(figsize = (8,18))
    subplot(4,1,1)
    plt.plot(Tarray, Darray)
    plt.xlabel("Time(s)")
    plt.ylabel("Diameter(cm)")
    plt.title("Diameter time domain change along position x = %d of original data" % position)
    subplot(4,1,2)
    plt.plot(Darrayfreqx, np.abs(Darrayfreqy))
    plt.ylabel("Amplitude")
    plt.xlabel("Frequency(Hz)")
    plt.title("Diameter frequency domain change along position x = %d via fourier transform" % position)
    subplot(4,1,3)
    plt.plot(Darrayfreqx, np.abs(Darrayfreqy))
    secondmax = sorted(np.abs(Darrayfreqy))[-2]
    plt.ylabel("Amplitude")
    plt.xlabel("Frequency(Hz)")
    plt.xlim([0,0.5])
    plt.ylim([0,secondmax*1.2])
    plt.title("Diameter zoomed in frequency domain change along position x = %d via fourier transform" % position)
    subplot(4,1,4)
    plt.plot(Tarray, DarrayInv)
    plt.xlabel("Time(s)")
    plt.ylabel("Diameter(cm)")
    plt.title("Diameter time domain change along position x = %d via inverse fourier transform" % position)
    plt.show()


# In[8]:


def DmapFreqMean(im_s, position):
    #calculate mean frequency at specific distance point over time from fourier transform
    #input: 2D image array: im_s
    #       the position index at distance map: position
    #output: mean frequency
    Tarray = np.linspace(0, shape(im_s)[0]/fps, num=shape(im_s)[0]) #time
    Darray = im_s[:,position] #get diameter at specific distance 
    Darray = ndimage.gaussian_filter1d(Darray, 50)  #smooth the curve
    Darrayfreqy = rfft(Darray)
    Darrayfreqx = rfftfreq(shape(im_s)[0],1/fps)  #perform fourier transform
    #calculate power 
    power = np.abs(Darrayfreqy) ** 2
    freq_mean = (sum(power*Darrayfreqx))/sum(power)
    return freq_mean


# In[9]:


def DmapFreqMode(im_s, position):
    #calculate mode frequency at specific distance point over time from fourier transform
    #mode frequency: frequency with highest amplitude
    #input: 2D image array: im_s
    #       the position index at distance map: position
    #output: mode frequency
    Tarray = np.linspace(0, shape(im_s)[0]/fps, num=shape(im_s)[0]) #time
    Darray = im_s[:,position] #get diameter at specific distance 
    Darray = ndimage.gaussian_filter1d(Darray, 50)  #smooth the curve
    Darrayfreqy = rfft(Darray)
    Darrayfreqx = rfftfreq(shape(im_s)[0],1/fps)  #perform fourier transform
    #find mode frequency
    maxindex = np.argsort(Darrayfreqy)[-2]  #get second max because the maximum is always at freq = 0 
    freq_mode = Darrayfreqx[maxindex]
    return freq_mode


# In[10]:


def RegionMeanFreqMode(im_s,Dmin = 0, Dmax = None, Tmin = 0, Tmax = None):
    #calculate the mean of the nonzero frequency mode
    #input: im_s: image array
    #       Dmin, Dmax, Tmin, Tmax: the region bbox to evaluate on
    
    if Dmax == None:
        Dmax = im_s.shape[1]/scale
    if Tmax == None:
        Tmax = im_s.shape[0]/fps
    im_s_evaluated = im_s[int(floor(Tmin*fps)):int(floor(Tmax*fps)),int(floor(Dmin*scale)):int(floor(Dmax*scale))]
    FreqMode = []
    for position in range(im_s_evaluated.shape[1]):
        FreqMode.append(DmapFreqMode(im_s_evaluated, position))
    return mean(FreqMode)


# In[11]:


def FeatureHeatmap(im_s):
    #plot heatmap for all features
    #input: im_s: raw diameter of image array
    feature_num = 2  #feature 0:mean frequency; feature 1: nonzero mode frequency
    slice_num = int(shape(im_s)[1])
    mtx = np.zeros((feature_num, slice_num))
    for i in range(slice_num):
        mtx[0,i] = DmapFreqMean(im_s,i)
        mtx[1,i] = DmapFreqMode(im_s,i)
    num_ticks = 10
    # the index of the position of yticks
    xticks = np.linspace(0, slice_num - 1, num_ticks, dtype=int)
    distance = np.linspace(0, shape(im_s)[1]/scale, num=shape(im_s)[1]) #distance
    xticklabels =  ["{:.2f}".format(distance[idx]) for idx in xticks]
    # the content of labels of these yticks
    plt.figure(figsize = (6,8))
    plt.subplot(2,1,1)
    ax1 = sns.heatmap(mtx[0,:].reshape(1,-1), cmap="YlGnBu",xticklabels=xticklabels,yticklabels = ["Mean Frequency"])
    ax1.set_xticks(xticks)
    ax1.set_xticklabels(xticklabels,rotation=20)
    ax1.set_xlabel("Distance(cm)")
    plt.subplot(2,1,2)
    ax2 = sns.heatmap(mtx[1,:].reshape(1,-1), cmap="YlGnBu",xticklabels=xticklabels,yticklabels = ["Mode Frequency"])
    ax2.set_xticks(xticks)
    ax2.set_xticklabels(xticklabels,rotation=20)
    ax2.set_xlabel("Distance(cm)")
    plt.show()
    return


# In[12]:


#find local minimum of STmap including on edges
def find_local_min(im_s, im_z, time_sigma = 5, spatial_sigma = 2, 
                   time_width = 50, time_prominence = [0.25,], peak_filter = None):
    #compute local minimum of a STmap in both row and column directions
    #first identify the local minimum along column (time) and along row (spatial) and find intersection
    #using padding and enable thresholding based on z score 
    #input: The raw image data array: im_s
    #       The z score image data array: im_z
    #       standard deviation for gaussian 1D filter in column local minimum identification: time_sigma, default = 5
    #       standard deviation for gaussian 1D filter in row local minimum identification: peak_filter, default = 2
    #       width parameter in identify peaks along column: time_width, default = 50
    #       prominence parameter in identify peaks along column: time_prominence, default = [0.25,]
    #       whether to use z score to filter peaks: peak_filter, default = None
    #output: The computed z score image data array: im_z_r
    #        The computed raw image data array: im_s_r
    #        row peaks(y coordinates): peak_x
    #        columnb peaks(x coordinates): peak_y
    
    #find local minimal in columns (time
    im_z_r = im_z.max()- im_z
    im_z_r = ndimage.gaussian_filter1d(im_z_r, sigma = time_sigma) #apply 1d guassian filter, sigma: standard deviation for Gaussian kernel
                                                                   #may be filter in individual direction?
    
    #perform padding around the image edge
    #padding tensors is adding zeros around the border of images to convert them to a shape that 
    #is amenable to the convolution operation without throwing away any pixel information
    orig_y, orig_x = im_z_r.shape
    pad_x = round(0.07 * im_z_r.shape[1])
    pad_x = pad_x if pad_x > 10 else 10  #set minimum pad x to be 10
    pad_y = round(0.07 * im_z_r.shape[0])
    pad_y = pad_y if pad_y > 100 else 100  #set minimum pad y to be 100
    im_z_r = np.pad(im_z_r, ((pad_y,pad_y),(pad_x,pad_x)), mode = "symmetric")
    #find local maximum along the array
    peakplot_c = np.zeros(im_z_r.shape)
    for i in range(im_z_r.shape[1]):
        peaks, _= scipy.signal.find_peaks(im_z_r[:, i], width = time_width, prominence = time_prominence) 
                    #The prominence of a peak measures how much the peak 
                    #stands out due to its intrinsic height and its location relative to other peaks
        for peak in peaks:
            peakplot_c[peak, i] = 1

    #find local minimal in rows
    im_s_r = im_s.max() - im_s
    if spatial_sigma != 0:
        im_s_r = ndimage.gaussian_filter1d(im_s_r, sigma = spatial_sigma)
    im_s_r = np.pad(im_s_r, ((pad_y,pad_y),(pad_x,pad_x)), mode = "symmetric")
    peakplot_r = np.zeros(im_s_r.shape)
    for i in range(im_s_r.shape[0]):
        peaks, _= scipy.signal.find_peaks(im_s_r[i, :])
        for peak in peaks:
            peakplot_r[i, peak] = 1
            
    #find the intersection between row and column minimum
    peakplot = peakplot_r * peakplot_c
    [peak_y, peak_x] = np.where(peakplot == 1)
    
    peak_x = peak_x - pad_x
    peak_y = peak_y - pad_y
    index = []
    #filter out the ones in added buffer edge
    for i in range(len(peak_y)): 
        if peak_x[i] < -1 or peak_x[i] >= orig_x or peak_y[i] < 0 or peak_y[i] >= orig_y:
            index += [i]
        if peak_x[i] == -1:
            peak_x[i] = 0
    index = np.array(index)
    peak_x = np.delete(peak_x, index)
    peak_y = np.delete(peak_y, index)
    # focus only on the region of interest (discard the added edges)
    im_z_r = im_z_r[pad_y:-pad_y, pad_x:-pad_x]
    im_s_r = im_s_r[pad_y:-pad_y, pad_x:-pad_x]
    
    #filter peaks based on z-score
    #the local maximum is defined as a peak if over certain threshold defined based on z score 
    if peak_filter != None:
        percentile = np.percentile(im_z_r, peak_filter, axis = 0 )
        index = []
        for i in range(len(peak_x)):
            if im_z_r[peak_y[i], peak_x[i]] < percentile[peak_x[i]]:
                index += [i]
        if len(index) != 0:
            index = np.array(index)
            peak_x = np.delete(peak_x, index)
            peak_y = np.delete(peak_y, index)
    return im_z_r, im_s_r, peak_x, peak_y
    


# In[13]:


def generate_combined_mask(im_z_r, peak_x, peak_y):
    #generate combined mask through flooding 
    #input: the local mimnimumly processed z score image array: im_z_r
    #       coordinates for peaks: peak_x, peak_y
    #output: the computed combined mask and visualized
    dot_queue = queue.Queue()
    dot_set = set()
    value = np.zeros(peak_x.shape)
    for i in range(peak_x.shape[0]):
        value[i] = im_z_r[peak_y[i], peak_x[i]]  #put the peak values in an array
                                                 
    index = np.argsort(-1*value) #get the descending index of sorted peak values
    for i in index:
        dot_queue.put((peak_x[i], peak_y[i]))  #put peak values in queue
        dot_set.add((peak_x[i], peak_y[i]))   #put peak values in set

    i = 0
    mask_c = np.zeros(im_z_r.shape)
    while len(dot_set) > 0:       
        seed = dot_queue.get()  #Remove and return an item from the queue
        if seed in dot_set:
            i += 1
            mask = oneSeed(seed, dot_set, im_z_r, 0.1, rate = 0.1)  #generate mask around the seed                            
            mask_c[mask & (mask_c == 0)] = i  #set the intersection of originally blank mask and wanted mask to be iteration number
                                            #why want iteration number i not 1?
    
    return mask_c


# In[14]:


def oneSeed(seed, dot_set, im_z, tol = 0.12, rate = 0.8):
    #generate flooding of one mask from one seed of peak point with combined information of other dots in mask
    #input: the flooding orgin peak point: seed
    #       the set of peak points: dot_set
    #       the z-score image array: im_z
    #       tolerance coefficient for flooding: tol (default: 0.12)
    #       rate coefficient for flooding: rate (default: 0.8)
    #output: the mask around the seed
    add_iter = 0
    #Mask corresponding to a flood fill.
    #Starting at a specific seed_point, connected points equal or within tolerance of the seed value are found.
    mask = flood(im_z, (seed[1],seed[0]), connectivity = 8, 
                 tolerance = tol*np.power(rate, add_iter)*im_z[seed[1],seed[0]]) #points with value within tolerance will be included in flooding area
    
    
    while True:
        added_dot = set()
        for dot in dot_set:
            #print("check:")
            #print(dot)
            if dotIsIn(dot[1], dot[0], mask): #if other dot is in mask, update the mask by adding the flood of those dots as well
                #print("add:")
                #print(dot)
                added_dot.add(dot)
                mask1 = flood(im_z, (dot[1], dot[0]), connectivity = 8, 
                              tolerance = tol*(im_z[seed[1],seed[0]]+
                              im_z[dot[1], dot[0]])/2)  #use new tolerance and flood origin to generate new masks
                mask = mask | mask1 #add mask 1 to mask
        for dot in added_dot:
            dot_set.remove(dot) #remove dot in dot set if added to the flooding of the mask
        add_iter += 1
        if len(added_dot) == 0:
            break       
    return mask


# In[15]:


def dotIsIn(dot_y, dot_x, mask):
    #check if dot is in the mask
    #input: y axis coordinate of dot: dot_y
    #       x axis coordinate of dot: dot_x
    #       mask to be checked: mask
    #output: return true if dot in mask, return false if dot not in mask
    return (mask[dot_y, dot_x] == 1)


# In[16]:


def annealRegions(mask_c, size_low_limit = 400):
    #anneal regions in mask and delete those under thershold
    #input: the mask to anneal: mask_c
    #       minimum size to remain after annealing: size_low_limit (default:400)
    #output: annealed mask mask_c
    properties = regionprops(mask_c.astype(int)) #measure the property of the region marked
    i = 1 #index of labeled region
    while True:
        #set solidity as standard for whether to anneal regions and termination
        #Ratio of pixels in the region to pixels of the convex hull image
        if properties[-i].solidity > 0.73 or i >= len(properties):
            break
        else:
            #delete regions that are too small (thresholded by solidity)
            mask_c[mask_c == (mask_c.max() - i + 1)] = 0
            print("mask ", mask_c.max() - i + 1, "= 0")
        i += 1
    
    #count regions to be removed
    to_remove = len(properties) - i + 1
    
    for i in np.arange(1, len(properties)):
        #print("threshold:", round(to_remove))
        if i <= to_remove: 
            mask_c = anneal(i, mask_c)
            #whatever, close the region using binary fill holes
            mask_c[binary_fill_holes(mask_c == i)] = i
            #delete regions that are too small after annealing
            if (mask_c == i).sum() < size_low_limit:
                #print("shape ",i, " is deleted because too small")
                mask_c[mask_c == i] = 0
    return mask_c


# In[17]:


def anneal(marker, mask_c):
    #anneal regions given mask and marker level (helper function in annealRegions())
    #input: anneal marker level: marker
    #       mask to be annealed: mask_c
    #output: annealed mask mask_c
    for j in np.arange(1, mask_c.max()):
        j_sum = (mask_c == j).sum()
        #print("j is ", j,"j sum is: ", j_sum)
        if (j != marker)& (j_sum > 0): 
            #print("try to anneal: ", marker, j)
            mask_j = mask_c == j  #get the entries where mask_c equal to level j
            mask_marker = mask_c == marker  #get the entries where mask_c equal to set standard marker
            #perform binary dilation
            #the locus of the points covered by the structuring element, when its center lies within the non-zero points of the image
            #get the overlapping region of two annealed masks
            result = (binary_dilation(mask_j, np.ones((3,3))) & binary_dilation(mask_marker, 
                                                                          np.ones((3,3)))).sum()
            #print("overlap result",marker, " and ", j," = ", result)
            if  result!= 0:
                print("annealing:", marker, " and ", j)
                mask_c[mask_marker] = 0
                mask_c[mask_j] = 0
                mask_c[mask_j | mask_marker] = marker
    return mask_c


# In[18]:


def filter_regions(mask_c, solidity_min = 0.5):
    #filter regions by using binary erosion and dilation
    #input: mask to be filtered: mask_c
    #       minimum solidity allowed after erosion and dilation: solidity_min
    #output: mask filtered: mask_c
    properties = regionprops(mask_c.astype(int))
    print("number of regions identified: ", len(properties))

    i = 0
    for prop in properties:
        i += 1
        # if convex shape, discard
        if mask_c[round(prop.centroid[0]), round(prop.centroid[1])] == 0:
            print("centroid zero:", i)
            mask_c[mask_c == i] = 0 

        elif prop.solidity < solidity_min:
            print("small solidity:",i)
            curr = mask_c == i
            footprint = np.ones((15,15))
            #perform erosion and dilation
            new_curr = binary_erosion(curr, np.ones((15, 15)))
            new_curr = binary_dilation(new_curr, np.ones((7,7)))
            curr_prop = regionprops(new_curr.astype(int))
            print(len(curr_prop))
            if new_curr.sum() < 0.2 * prop.area:
                print("remove", i, "due to diminish to zero")
                mask_c[mask_c == i] = 0  
            elif curr_prop[0].solidity < prop.solidity:
                print("remove ", i, "due to small solidity")
                mask_c[mask_c == i] = 0
            else:
                mask_c[mask_c == i] = 0
                mask_c[new_curr & (mask_c == 0)] = i #set the intersection of two masks

    properties = regionprops(mask_c.astype(int))     
    print("number of regions left: ",len(properties))
    return mask_c


# In[19]:


def clean_mask_c(mask_c):
    #clean mask
    #input: mask mask_c
    #output: cleaned version mask_c
    i = 1
    for j in np.arange(1, mask_c.max()+1):
        mask_curr = mask_c == j
        if mask_curr.sum() > 0: #if occurence of level j is nonzero
            mask_c[mask_curr] = i
            i += 1    #reassign value
    #plt.close("all")
    plt.figure()
    plt.imshow(mask_c, aspect = aspect, extent=[0,shape(mask_c)[1]/scale,shape(mask_c)[0]/fps,0])
    #add annotation to each region
    properties = regionprops(mask_c.astype(int))
    for i in np.arange(1, mask_c.max()+1):
        mask_curr = mask_c == i
        coord = [properties[int(i-1)].centroid[1]/scale,properties[int(i-1)].centroid[0]/fps]
        plt.annotate(str(int(i)),coord,color='white',weight="bold")
    plt.show()
    return mask_c


# In[20]:


def clean_mask_c_GUI(mask_c, x_inch, y_inch, dpi, file_prefix):
    #clean mask for GUI
    #input: mask mask_c
    #       x_inch: x size of the mask
    #       y_inch: y size of the mask
    #       dpi: resolution
    #       file_prefix: the filename
    i = 1
    for j in np.arange(1, mask_c.max()+1):
        mask_curr = mask_c == j
        if mask_curr.sum() > 0: #if occurence of level j is nonzero
            mask_c[mask_curr] = i
            i += 1    #reassign value
    fig = plt.figure(frameon=False, figsize=(x_inch, y_inch), dpi=dpi)
    ax  = plt.Axes(fig, [0., 0., 1., 1.])
    ax.set_axis_off()
    fig.add_axes(ax)
    ax.imshow(mask_c, aspect="auto")
    fig.savefig(f"{file_prefix}_reindexed.png", dpi=dpi)
    return mask_c


# In[21]:


def regionEllipsePropertyDisp(mask_c, im_z):
    #display region properties on mask based on the assumption that the regions are ellpises
    #input: mask to be analyzed: mask_c
    #       image array im_z
    properties = regionprops(mask_c.astype(int))
    i = 0
    for prop in properties:
        i = i+1
        print(i)
        print("next one:")
        print("area:", prop.area)
        print("bbox", prop.bbox)
        print("centroid z depth:", im_z[round(prop.centroid[0]), round(prop.centroid[1])])
        print("area_convex", prop.area_convex)
        print("area_filled:", prop.area_filled)
        print("axis_major_length", prop.axis_major_length)
        print("axis_minor_length", prop.axis_minor_length)
        print("centroid:", prop.centroid)
        print("eccentricity:", prop.eccentricity)
        print("orientation", prop.orientation)
        print("perimeter/area", prop.perimeter/prop.area)
        print("solidity", prop.solidity)


# In[22]:


def edgeDetection_GUI(mask_c, x_inch, y_inch, dpi, file_prefix, threshold1=10, threshold2=150, FigDisplay=False):
    #edge detection for all labled region
    #input:mask_c: mask with labeled region
    #       x_inch: x size of the mask
    #       y_inch: y size of the mask
    #       dpi: resolution
    #       file_prefix: the filename
    #      threshold1:lower threshold for sobel edge detection(default: 100)
    #      threshold2:upper threshold for sobel edge detection(default: 150)
    #output: visualization for all edge detection
    #        return all edge edges_all
    mask_curr_all = mask_c != 0  #get all nonzero region
    mask_curr_all = mask_curr_all.astype(float)
    img_blur_all = cv2.GaussianBlur(mask_curr_all, (3,3), 0)
    sobelx    = cv2.Sobel(src=img_blur_all, ddepth=cv2.CV_64F, dx=1, dy=0, ksize=5) # Sobel Edge Detection on the X axis
    sobely    = cv2.Sobel(src=img_blur_all, ddepth=cv2.CV_64F, dx=0, dy=1, ksize=5) # Sobel Edge Detection on the Y axis
    sobelxy   = cv2.Sobel(src=img_blur_all, ddepth=cv2.CV_64F, dx=1, dy=1, ksize=5) # Combined X and Y Sobel Edge Detection
    edges_all = cv2.Canny(image=(img_blur_all*255).astype(np.uint8), 
                          threshold1=threshold1, threshold2=threshold2)             # Canny Edge Detection

    if display:
        fig = plt.figure(frameon=False, figsize=(x_inch, y_inch), dpi=dpi)
        ax  = plt.Axes(fig, [0., 0., 1., 1.])
        ax.set_axis_off()
        fig.add_axes(ax)
        ax.imshow(sobelx   , aspect='auto')
        ax.imshow(sobely   , aspect='auto')
        ax.imshow(sobelxy  , aspect='auto')
        ax.imshow(edges_all, aspect='auto', cmap="binary")
        fig.savefig(f"{file_prefix}_updated.png", dpi=dpi)

    output1 = np.maximum(sobelx, sobely)
    output2 = np.maximum(sobelxy, edges_all)
    overall = np.maximum(output1, output2)
    return edges_all


# In[23]:


def edgeDetectionIndividual_GUI(mask_c,regionIndex, threshold1=100, threshold2=150,FigDisplay = False):
    #edge detection for one specific labled region for GUI
    #input:mask_c: mask with labeled region
    #      regionIndex: region index label for mask
    #      threshold1:lower threshold for sobel edge detection(default: 100)
    #      threshold2:upper threshold for sobel edge detection(default: 150)
    #      FigDisplay:whether to display figure and save
    #output: visualization for all edge detection
    #        return all edge edges_all
    mask_curr = mask_c == regionIndex  #grab specific region
    mask_curr = mask_curr.astype(float)
    img_blur = cv2.GaussianBlur(mask_curr, (3,3), 0)
    sobelx = cv2.Sobel(src=img_blur, ddepth=cv2.CV_64F, dx=1, dy=0, ksize=5) # Sobel Edge Detection on the X axis
    sobely = cv2.Sobel(src=img_blur, ddepth=cv2.CV_64F, dx=0, dy=1, ksize=5) # Sobel Edge Detection on the Y axis
    sobelxy = cv2.Sobel(src=img_blur, ddepth=cv2.CV_64F, dx=1, dy=1, ksize=5) # Combined X and Y Sobel Edge Detection
    edges = cv2.Canny(image=(img_blur*255).astype(np.uint8), threshold1 = threshold1, threshold2= threshold2) # Canny Edge Detection   
    if FigDisplay == True:
        fig = plt.figure()
        ax  = plt.Axes(fig, [0., 0., 1., 1.])
        fig.add_axes(ax)
        ax.imshow(sobelx   , aspect='auto')
        ax.imshow(sobely   , aspect='auto')
        ax.imshow(sobelxy  , aspect='auto')
        ax.imshow(edges, aspect='auto', cmap="binary")
        fig.savefig(f"{file_prefix}_region{regionIndex}_edge.png")
    return edges


# In[24]:


def EdgeIndividualPivot_GUI(mask_c,regionIndex,FigDisplay = False):
    #pivot the edges of individual contraction region to be the origin for start point for GUI
    #input: mask_c: mask generated
    #       regionIndex: region investigated
    #       FigDisplay:whether to display figure and save
    #output: X: 2D array with coordinates for repivoted edges
    edges = edgeDetectionIndividual_GUI(mask_c,regionIndex, FigDisplay = False)
    #detect wave front in selected region
    start_f = 0
    waveedgex = []
    waveedgey = []
    for i in range(edges.shape[0]):  #row index/time index
            if sum(edges[i]) != 0: #when there is none zero index along the same time index
                for j in range(edges.shape[1]): #column index/distance index
                        if edges[i][j] != 0:  #assume leftmost leading front is the start point
                            if start_f == 0:
                                start_time = i
                                start_distance  = j
                                waveedgex.append(0) #the x coordinate is distance
                                waveedgey.append(0)  #the y coordinate is time
                                start_f = 1 
                            else:
                                time_val = (i - start_time)/fps
                                distance_val = (j - start_distance)/scale
                                waveedgex.append(distance_val)
                                waveedgey.append(time_val)
    X = []
    for i in range(len(waveedgex)):
        X.append([waveedgex[i], waveedgey[i]])   #note that X returned is sorted by y value
    if FigDisplay == True:
        waveedgey = np.array(waveedgey)
        waveedgey = max(waveedgey) - waveedgey
        fig = plt.figure()
        plt.title("Wave edge repivoted for contraction region label = {}".format(regionIndex))
        plt.scatter(waveedgex,waveedgey)
        plt.annotate("wave start point",(0,max(waveedgey)),color='black',weight="bold")
        plt.ylabel("Pivoted Time(s)")
        plt.xlabel("Pivoted Distance(s)")
        fig.savefig(f"{file_prefix}_region{regionIndex}_repivoted.png")
    return [X,start_time,start_distance]


# In[25]:


def getpropagationDir_GUI(mask_c,regionIndex,FigDisplay = False):
    #determine the propagation direction of the region for GUI
    #either be anterograde, retrograde, bidirectional propagation
    #input: mask_c: mask generated
    #       regionIndex: region investigated
    #      FigDisplay:whether to display figure and save
    #output: propagation_dir: type of propagation
    [X,start_time,start_distance] = EdgeIndividualPivot_GUI(mask_c,regionIndex,FigDisplay = FigDisplay)
    X = np.array(sorted(X))  #sorted in ascending order
    if X[0,0] == 0:
        propagation_dir = "anterograde propagation"
    elif X[-1,0] == 0:
        propagation_dir = "retrograde propagation"
    else:
        propagation_dir = "bidirectional propagation"
    return propagation_dir


# In[26]:


def AnterogradeFront(X,tolerance_intercept = 0.1, tolerance_slope = 1000):
    #wavefront detection for anterograde propagation
    #input: X: 2D array for pivoted edge
    #      tolerance_intercept: parameter for neighborhood detection mode (initial boost from zero)
    #      tolerance_slope: parameter for neighborhood detection mode (subsequent slope)
    #output: wavefront: 2D array for wavefront detected 
    wavefront = []
    xpivot = 0
    ypivot = 0
    for i in range(X.shape[0]):  #go from left to right
        if ypivot == 0:
            if abs(X[i,1]) < tolerance_intercept: #initial neighborhood around origin
                wavefront.append([X[i,0],X[i,1]])
                xpivot = X[i,0]
                ypivot = X[i,1]     
        else:  #include if bounded in predefined box
            if (X[i,1] > ypivot - (X[i,0] - xpivot)*tolerance_slope) & (X[i,1] < ypivot + (X[i,0] - xpivot)*tolerance_slope):
                wavefront.append([X[i,0],X[i,1]])
                xpivot = X[i,0]
                ypivot = X[i,1] 
    return wavefront


# In[27]:


def RetrogradeFront(X,tolerance_intercept = 0.1, tolerance_slope = 1000):
    #wavefront detection for retrograde propagation
    #input: X: 2D array for pivoted edge
    #      tolerance_intercept: parameter for neighborhood detection mode (initial boost from zero)
    #      tolerance_slope: parameter for neighborhood detection mode (subsequent slope)
    #output: wavefront: 2D array for wavefront detected 
    wavefront = []
    xpivot = 0
    ypivot = 0
    for i in range(X.shape[0]-1, -1, -1):  #go from right to left
        if ypivot == 0:
            if abs(X[i,1]) < tolerance_intercept: #initial neighborhood around origin
                wavefront.append([X[i,0],X[i,1]])
                xpivot = X[i,0]
                ypivot = X[i,1]     
        else: #include if bounded in predefined box
            if (X[i,1] < ypivot - (X[i,0] - xpivot)*tolerance_slope) & (X[i,1] > ypivot + (X[i,0] - xpivot)*tolerance_slope):
                wavefront.append([X[i,0],X[i,1]])
                xpivot = X[i,0]
                ypivot = X[i,1] 
    return wavefront


# In[28]:


def waveFrontDetection_GUI(mask_c,regionIndex,propagation_dir, Detectmode = "DBSCAN", anterograde_tolerance_intercept = 0.1, 
                       anterograde_tolerance_slope = 1000, retrograde_tolerance_intercept = 0.1, 
                       retrograde_tolerance_slope = 1000, FigDisplay = False):
    #edge detection for one specific region
    #input:mask_c: mask with labeled region
    #      regionIndex: region index label for mask
    #      propagation_dir: the type of propagation based on direction
    #      detectMode: choose from two modes: DBSCAN unsupervised clustering OR neighborhood detection
    #      anterograde_tolerance_intercept: parameter for neighborhood detection mode (initial boost from zero)
    #      anterograde_tolerance_slope: parameter for neighborhood detection mode (subsequent slope)
    #      retrograde_tolerance_intercept: parameter for neighborhood detection mode (initial boost from zero)
    #      retrograde_tolerance_slope: parameter for neighborhood detection mode (subsequent slope)
    #      FigDisplay:whether to display figure and save
    #output: visualization for wavefront and waveback detection, wavefront detection
    #        return coordinate points for wavefront  
    #sanity check
    if (Detectmode != "DBSCAN") & (Detectmode != "neighborhood"):
        print("No method found!")
        return

    if Detectmode == "DBSCAN":
        [X,start_time,start_distance] = EdgeIndividualPivot_GUI(mask_c,regionIndex,FigDisplay = False)
        #run unsupervised classification for wavefront and waveback
        X = np.array(X)
        clustering = DBSCAN(eps=3, min_samples=2).fit(X)
        if (len(np.unique(clustering.labels_)) == 2) &(propagation_dir != "bidirectional propagation"):  #check if DBSCAN generates two clusters as intended
            #detect wavefront
            class0_ind = []
            class1_ind = []
            for i in range(len(clustering.labels_)):
                if clustering.labels_[i] == 0:
                    class0_ind.append(i)
                else:
                    class1_ind.append(i)
            if X[class0_ind[0]][1] < X[class1_ind[0]][1]:
                wavefront_class = class0_ind
            else:
                wavefront_class = class1_ind
            wavefront = np.array([X[i] for i in wavefront_class])
        else:  #if DBSCAN failed, change to neighborhood method
            Detectmode = "neighborhood"  #change detection method
            
    if Detectmode == "neighborhood":
        [X,start_time,start_distance] = EdgeIndividualPivot_GUI(mask_c,regionIndex,FigDisplay = False)
        X = np.array(sorted(X))
        X_neg = []
        X_pos = []
        for i in range(X.shape[0]):
            if X[i,0] <= 0:
                X_neg.append([X[i,0],X[i,1]]) #get backward propagation part
            if X[i,0] >= 0:
                X_pos.append([X[i,0],X[i,1]]) #get forward propagation part
        X_pos = np.array(X_pos)
        X_neg = np.array(X_neg)
        #forward wavefront
        if propagation_dir == "anterograde propagation":
            wavefront =  AnterogradeFront(X_pos,tolerance_intercept = anterograde_tolerance_intercept, tolerance_slope = anterograde_tolerance_slope)
        elif propagation_dir == "retrograde propagation":
            wavefront =  RetrogradeFront(X_neg,tolerance_intercept = retrograde_tolerance_intercept, tolerance_slope = retrograde_tolerance_slope)
        elif propagation_dir == "bidirectional propagation":
            
            Forwardwavefront = AnterogradeFront(X_pos,tolerance_intercept = anterograde_tolerance_intercept, tolerance_slope = anterograde_tolerance_slope)
            #backward wavefront
            Backwardwavefront = RetrogradeFront(X_neg,tolerance_intercept = retrograde_tolerance_intercept, tolerance_slope = retrograde_tolerance_slope)

            #check if any forward or backward waves are void
            if len(Backwardwavefront) == 0:
                propagation_dir = "anterograde propagation"
            if len(Forwardwavefront) == 0:
                propagation_dir = "retrograde propagation"
            if (len(Backwardwavefront) == 0) & (len(Forwardwavefront) == 0):
                print("Problematic contraction without propagation detection!")
                return
            
            Backlen = len(Backwardwavefront)
            Backwardwavefront.extend(Forwardwavefront)
            wavefront = Backwardwavefront
            
    #for all cases:
    wavefront = np.array(wavefront)
    wavefront[:,1] = max(wavefront[:,1]) - wavefront[:,1]
    if FigDisplay == True:

        fig = plt.figure()
        plt.title("wavefront detected for contraction region label = {}".format(regionIndex))
        plt.scatter(wavefront[:,0],wavefront[:,1])
        plt.ylabel("Time(s)")
        plt.xlabel("Distance(cm)")
        fig.savefig(f"{file_prefix}_region{regionIndex}_wavefront.png")

    #determine return type
    if propagation_dir == "bidirectional propagation":
        print(type(wavefront))
        return [wavefront,Backlen]
    else:
        print(type(wavefront))
        return [wavefront]




# In[29]:


def sigmoid(x, L ,x0, k, b):
    #define sigmoid curve for fitting of wavefront
    #input: x: x coordinate
    #       L:maximum point
    #       x0: mid point for 50% maximum
    #       k: slope for sigmoid curve
    #       b: minimum point
    #output: y: y coordinate of function
    y = L / (1 + np.exp(-k*(x-x0))) + b
    return (y)


# In[30]:


def sigmoidWaveFrontFit_GUI(wavefront,regionIndex,dirFlag, FigDisplay = False):
    #add padded area on the left and right of the curve for GUI
    #input: wavefront: wavefront edge
    #       regionIndex: region index label for mask
    #       dirFlag: specify propagation direction
    #       FigDisplay:whether to display figure and save
    #output: the visualization of sigmoidal curve fit
    #        popt: the optimized parameter for sigmoid curve:[L, x0, k, b]
    etd_len = int(shape(wavefront)[0]*0.5)
    wavefront_extended = zeros([shape(wavefront)[0]+etd_len*2,shape(wavefront)[1]])
    wavefront_extended[:etd_len,1] = wavefront[0,1]
    wavefront_extended[-etd_len:,1] = wavefront[-1,1]
    if wavefront[-1,0] > 0:  #forward propagation
        wavefront_extended[:etd_len,0] = wavefront[0,0] - np.linspace(0.01*etd_len,0.01,etd_len)
        wavefront_extended[-etd_len:,0] = wavefront[-1,0] + np.linspace(0.01, 0.01*etd_len,etd_len)
    else:   #backward propagation
        wavefront_extended[:etd_len,0] = wavefront[0,0] + np.linspace(0.01, 0.01*etd_len,etd_len)  
        wavefront_extended[-etd_len:,0] = wavefront[-1,0] - np.linspace(0.01*etd_len,0.01,etd_len)

    wavefront_extended[etd_len:-etd_len,:] = wavefront
    xdata = wavefront_extended[:,0]
    ydata = wavefront_extended[:,1]

    p0 = [max(ydata), np.median(xdata),1,min(ydata)] # this is an mandatory initial guess
    popt, pcov = curve_fit(sigmoid, xdata, ydata,p0, method='dogbox')
    x = np.linspace(min(xdata), max(xdata), 100)
    y = sigmoid(x, *popt)
    if FigDisplay == True:
        fig = plt.figure()
        plt.plot(xdata, ydata, 'o', label='data')
        plt.plot(x,y, label='fit')
        plt.title("Fitted {} wavefront for contraction region label = {}".format(dirFlag, regionIndex))
        plt.xlabel("Distance(cm)")
        plt.ylabel("Time(s)")
        plt.legend(loc='best')
        fig.savefig(f"{file_prefix}_region{regionIndex}_{dirFlag}_wavefrontSigmoidFit.png")
    return popt


# In[31]:


def propregionFeature_GUI(mask_c,im_z, im_s, regionIndex, contractionType, propagation_dir, Detectmode = "DBSCAN", 
                  anterograde_tolerance_intercept = 0.1, anterograde_tolerance_slope = 1000, 
                  retrograde_tolerance_intercept = 0.1, retrograde_tolerance_slope = 1000, FigDisplay = False):
    #this funciton is for GUI
    #display features of the region in contraction type under propagation family: propagation & interrupted small contraction
    #input: mask_c: mask with labeled region
    #       im_z: z score STmap
    #       regionIndex: region index label for mask
    #       contractionType: choose from two types of contraction: propagation OR interuppted small contraction
    #       propagation_dir: propagation direction label: anterograde propagation OR retrograde propagation OR bidirectional propagation    
    #       detectMode: choose from two wavefront detection modes: DBSCAN unsupervised clustering OR neighborhood detection
    #       tolerance_intercept: parameter for neighborhood detection mode (initial boost from zero)
    #       tolerance_slope: parameter for neighborhood detection mode (subsequent slope)
    #      FigDisplay:whether to display figure and save
    #output:featuredict: a dictionary of region properties
    if (contractionType == "propagation") or (contractionType == "interrupted small contraction"):       
        if propagation_dir == "anterograde propagation":
            [wavefront] = waveFrontDetection_GUI(mask_c,regionIndex,propagation_dir, Detectmode = "DBSCAN", 
                                                 anterograde_tolerance_intercept = anterograde_tolerance_intercept, 
                                                 anterograde_tolerance_slope = anterograde_tolerance_slope, 
                                                 retrograde_tolerance_intercept = retrograde_tolerance_intercept, 
                                                 retrograde_tolerance_slope = retrograde_tolerance_slope, 
                                                 FigDisplay = FigDisplay)
            featuredict = {}
            featuredict["sample name"] = file_prefix
            dirFlag = "anterograde"
            featuredict["class"] = f"{contractionType[:11]}_{dirFlag}"
            featuredict["region index"] = f"region {regionIndex}"
            featuredict["anterograde tolerance intercept"] = anterograde_tolerance_intercept
            featuredict["anterograde tolerance slope"] = anterograde_tolerance_slope
            featuredict["retrograde tolerance intercept"] = retrograde_tolerance_intercept
            featuredict["retrograde tolerance slope"] = retrograde_tolerance_slope
            featuredict = waveFrontFeature_GUI(featuredict,mask_c,regionIndex,wavefront,contractionType,dirFlag,FigDisplay)
            featuredict = basicRegionFeature(featuredict,mask_c, im_z,im_s, regionIndex)
            return featuredict
        elif propagation_dir == "retrograde propagation":
            [wavefront] = waveFrontDetection_GUI(mask_c,regionIndex,propagation_dir, Detectmode = "DBSCAN", 
                                                 anterograde_tolerance_intercept = anterograde_tolerance_intercept, 
                                                 anterograde_tolerance_slope = anterograde_tolerance_slope, 
                                                 retrograde_tolerance_intercept = retrograde_tolerance_intercept, 
                                                 retrograde_tolerance_slope = retrograde_tolerance_slope, 
                                                 FigDisplay = FigDisplay)
            featuredict = {}
            featuredict["sample name"] = file_prefix
            dirFlag = "retrograde"
            featuredict["class"] = f"{contractionType[:11]}_{dirFlag}"
            featuredict["region index"] = f"region {regionIndex}"
            featuredict["anterograde tolerance intercept"] = anterograde_tolerance_intercept
            featuredict["anterograde tolerance slope"] = anterograde_tolerance_slope
            featuredict["retrograde tolerance intercept"] = retrograde_tolerance_intercept
            featuredict["retrograde tolerance slope"] = retrograde_tolerance_slope
            featuredict = waveFrontFeature_GUI(featuredict,mask_c,regionIndex,wavefront,contractionType,dirFlag, FigDisplay)
            featuredict = basicRegionFeature(featuredict,mask_c, im_z, im_s,regionIndex)
            return featuredict
        elif propagation_dir == "bidirectional propagation":
            [wavefront, Backlen] = waveFrontDetection_GUI(mask_c,regionIndex,propagation_dir, Detectmode = "DBSCAN", 
                                                 anterograde_tolerance_intercept = anterograde_tolerance_intercept, 
                                                 anterograde_tolerance_slope = anterograde_tolerance_slope, 
                                                 retrograde_tolerance_intercept = retrograde_tolerance_intercept, 
                                                 retrograde_tolerance_slope = retrograde_tolerance_slope, 
                                                 FigDisplay = FigDisplay)
            
            Forwardwavefront = wavefront[Backlen:]
            Backwardwavefront = wavefront[:Backlen]
            featuredict = {}
            featuredict["sample name"] = file_prefix
            featuredict["class"] = f"{contractionType[:11]}_bidirectional"
            featuredict["region index"] = f"region {regionIndex}"
            featuredict["anterograde tolerance intercept"] = anterograde_tolerance_intercept
            featuredict["anterograde tolerance slope"] = anterograde_tolerance_slope
            featuredict["retrograde tolerance intercept"] = retrograde_tolerance_intercept
            featuredict["retrograde tolerance slope"] = retrograde_tolerance_slope
            dirFlag = "anterograde"
            featuredict = waveFrontFeature_GUI(featuredict,mask_c,regionIndex,Forwardwavefront,contractionType,dirFlag,FigDisplay)
            dirFlag = "retrograde"
            featuredict = waveFrontFeature_GUI(featuredict,mask_c,regionIndex,Backwardwavefront,contractionType,dirFlag,FigDisplay)
            featuredict = basicRegionFeature(featuredict,mask_c, im_z,im_s, regionIndex)

            return featuredict
    else:
        print("Contraction type not found!")


# In[32]:


def waveFrontFeature_GUI(featuredict,mask_c,regionIndex,wavefront,contractionType,dirFlag, FigDisplay = False):
    #extract features for wave front with fitted sigmoid curve
    #input: featuredict:dictionary to put features
    #       mask_c: mask generated
    #       regionIndex: index of region
    #       popt: fitted sigmoid curve
    #       wavefront: wavefront array
    #       contractionType: propagation OR interrupted contraction region
    #       dirFlag: anterograde OR retrograde
    #       FigDisplay:whether to display figure and save
    #output: print the properties
    #        return a list of property
    [X,start_time,start_distance]= EdgeIndividualPivot_GUI(mask_c,regionIndex,FigDisplay = FigDisplay)
    #basic scenario without fitting sigmoid curve
    #calculate propagation start distance percentage position
    propagation_start = start_distance/(mask_c.shape[1])
    print("The", dirFlag,"propagation start point position percentage is:", propagation_start*100,"%")
    #calculate the propgation end distance percentage position
    propagation_end = (start_distance/scale + wavefront[-1,0])/(mask_c.shape[1]/scale)
    print("The", dirFlag,"propagation end point position percentage is:", propagation_end*100,"%")
    #calculate the propagation distance span
    propagation_Dspan = max(wavefront[:,0]) - min(wavefront[:,0])
    print("The", dirFlag,"propagation distance for the contraction is:", propagation_Dspan,"cm")
    #calculate the propagation distance percentage
    propagation_Dper = propagation_Dspan/(mask_c.shape[1]/scale)
    print("The", dirFlag,"propagation distance percentage is:",propagation_Dper*100,"%")
    #calculate the propagation time span
    propagation_Tspan = max(wavefront[:,1]) - min(wavefront[:,1])
    print("The", dirFlag,"propagation time for the contraction is:", propagation_Tspan,"s")
    #calculate overall velocity
    overall_velocity = (max(wavefront[:,0]) - min(wavefront[:,0]))/(max(wavefront[:,1]) - min(wavefront[:,1]))
    print("The overall", dirFlag,"propagation velocity for the contraction is:", overall_velocity,"cm/s")
    
    if dirFlag == "anterograde":
        featuredict["anterograde propagation start point percentage(%)"] = propagation_start*100
        featuredict["anterograde propagation end point percentage(%)"] = propagation_end*100
        featuredict["anterograde propagation distance span(cm)"] = propagation_Dspan
        featuredict["anterograde propagation distance percentage(%)"] = propagation_Dper*100
        featuredict["anterograde propagation time span(s)"] = propagation_Tspan
        featuredict["overall anterograde velocity(cm/s)"] =overall_velocity 
    elif dirFlag == "retrograde":
        featuredict["retrograde propagation start point percentage(%)"] = propagation_start*100
        featuredict["retrograde propagation end point percentage(%)"] = propagation_end*100
        featuredict["retrograde propagation distance span(cm)"] = propagation_Dspan
        featuredict["retrograde propagation distance percentage(%)"] = propagation_Dper*100
        featuredict["retrograde propagation time span(s)"] = propagation_Tspan
        featuredict["overall retrograde velocity(cm/s)"] =overall_velocity 

    if contractionType == "propagation":
        popt = sigmoidWaveFrontFit_GUI(wavefront,regionIndex,dirFlag,FigDisplay)
        #calculate the propagation front slope mid point percentage position
        propagation_mid = (start_distance/scale + popt[1])/(mask_c.shape[1]/scale)
        print("The", dirFlag,"propagation mid point position percentage is:", propagation_mid*100,"%")
        #calculate the propagation front mid point slope
        propagation_midslope = popt[2]
        print("The", dirFlag,"propagation mid point slope is:", propagation_midslope)
        
        if dirFlag == "anterograde":
            featuredict["anterograde propagation max velocity position percentage(%)"] = propagation_mid*100
            featuredict["anterograde propagation max velocity(cm/s)"] = propagation_midslope
        elif dirFlag == "retrograde":
            featuredict["retrograde propagation max velocity position percentage(%)"] = propagation_mid*100
            featuredict["retrograde propagation max velocity(cm/s)"] = propagation_midslope

    return featuredict


# In[33]:


def basicRegionFeature(featuredict,mask_c, im_z, im_s, regionIndex):
    #extract features for wave front
    #input: featuredict: dictionary to put features
    #       mask_c: mask generated
    #       im_z: z score image array
    #       regionIndex: index of region
    #output: print the properties
    #        return a list of property
    #other region properties relate to mask
    properties = regionprops(mask_c.astype(int))
    prop = properties[regionIndex-1]
    contraction_area = prop.area/(scale*scale)
    print("Contraction region area:", contraction_area,"cm^2")

    solidity = prop.solidity
    print("Contraction region solidity: ", solidity)
    centroid_time = prop.centroid[0]/fps
    centroid_distance_per = prop.centroid[1]/im_z.shape[1]
    print("Contraction centroid time position:", centroid_time,"s, distance position percentage:",centroid_distance_per*100 ,"%")

    #other region properties relate to intensity of z score STmap
    im_z_mask = im_z.copy()
    im_z_mask_lin = im_z_mask[mask_c == regionIndex]
    intensity_max = im_z_mask_lin.min()
    intensity_mean = im_z_mask_lin.mean()
    #added for raw diameter
    im_s_mask = im_s.copy()
    im_s_mask_lin = im_s_mask[mask_c == regionIndex]
    intensity_max_raw = im_s_mask_lin.min()
    intensity_mean_raw = im_s_mask_lin.mean()
    
    im_z_mask[mask_c != regionIndex] = 0
    intensity_max_coor_pixel = np.argwhere(im_z_mask == intensity_max)
    intensity_max_coor_TD = intensity_max_coor_pixel.astype(float)
    intensity_max_coor_TD[:,0] = intensity_max_coor_pixel[:,0]/fps
    intensity_max_coor_TD[:,1] = intensity_max_coor_pixel[:,1]/scale    
    
    print("Maximum contraction intensity:",intensity_max)
    print("Raw diameter at maximum contraction intensity:", intensity_max_raw)
    print("Pixel coordinate of maximum contraction intensity is:",intensity_max_coor_pixel)
    print("Time distance coordinate of maximum contraction intensity is:", intensity_max_coor_TD)
    print("Mean contraction intensity:",intensity_mean)
    print("Raw diameter at mean contraction intensity:", intensity_mean_raw)

    
    #write to dict
    featuredict["contraction area(cm^2)"] = contraction_area
    featuredict["solidity"] = solidity
    featuredict["maximum z score contraction intensity"] = intensity_max
    featuredict["pixel coordinate of max z score contraction intensity"] = intensity_max_coor_pixel
    featuredict["time distance coordinate of max z score contraction intensity"] = intensity_max_coor_TD
    featuredict["mean z score contraction intensity"] = intensity_mean
    featuredict["time at contraction centroid(s)"] = centroid_time
    featuredict["distance percentage at contraction centroid(%)"] = centroid_distance_per*100
    featuredict["raw diameter at maximum contraction"] = intensity_max_raw
    featuredict["raw diameter at mean contraction"] = intensity_mean_raw
    
    return featuredict


# In[34]:


def rippleregionFeature(mask_c,im_s, im_z,regionIndex,Dmin = 0, Dmax = None, Tmin = 0, Tmax = None):
    #extract ripple features from specified regions
    #input: mask mask_c
    #       raw diameter STmap im_s
    #       z score STmap im_z
    #       region index regionIndex
    #       distance left bound: Dmin
    #       distance right bound: Dmax
    #       Time left bound: Tmin
    #       Time right bound: Tmax
    #output: feature dictionary
    if Dmax == None:
        Dmax = im_s.shape[1]/scale
    if Tmax == None:
        Tmax = im_s.shape[0]/fps
    featuredict = {}
    Dstartper = Dmin/(im_s.shape[1]/scale)
    print("The ripple region start point distance percentage is:",Dstartper*100,"%")
    Dendper = Dmax/(im_s.shape[1]/scale)
    print("The ripple region end point distance percentage is:",Dendper*100,"%")  
    Dspan = Dmax - Dmin
    print("The distance span of ripple region is:",Dspan,"cm")
    Tspan = Tmax - Tmin
    print("The time span of ripple region is:",Tspan,"s")
    FreqMode = RegionMeanFreqMode(im_s,Dmin, Dmax, Tmin, Tmax)
    print("The mean of nonzero mode frequency is:",FreqMode)   
    featuredict["sample name"] = sampleIndex
    featuredict["region index"] = f"region {regionIndex}"
    featuredict["class"] = "ripple"
    featuredict["distance start position percentage(%)"] = Dstartper*100
    featuredict["distance end position percentage(%)"] = Dendper*100
    featuredict["distance span(cm)"] = Dspan
    featuredict["time span(s)"] = Tspan
    featuredict["mean of mode ripple frequency(Hz)"] = FreqMode
    featuredict = basicRegionFeature(featuredict,mask_c, im_z,im_s, regionIndex)
    return featuredict


# In[ ]:





# In[35]:


def write_all_to_results_GUI(sample_featuredict,filename):
    #write results to existing excel that stores all results
    #input: finalized sample_featuredict
    #       excel filename to write final data to
    ExcelWorkbook = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine = 'openpyxl')
    writer.book = ExcelWorkbook
    #write each record to excel
    for regionLabel,featuredict in sample_featuredict.items():
        featuredf = pd.DataFrame(featuredict, index=[0])
        sheetname = featuredf.loc[0]["class"]
        featuredf.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    writer.save()
    writer.close()


# In[ ]:





# In[36]:


get_ipython().system('jupyter nbconvert --to script GUI_Helper_fcn.ipynb')


# In[ ]:





# In[ ]:





# In[37]:


{i:0 for i in range(5)}


# In[ ]:




